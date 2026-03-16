from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
import os

from utils import is_unreachable_screenshot  # Import OCR check function

def write_results_to_excel(results, output_path="results.xlsx"):
    """
    Writes results to Excel with URL, Ports, and embedded Screenshots.
    Skips unreachable sites based on screenshot content.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Screenshots"

    # Headers
    ws['A1'] = "URL"
    ws['B1'] = "Open Ports"
    ws['C1'] = "Screenshot"

    row = 2
    for result in results:
        img_path = result.get("screenshot")
        
        # Skip if unreachable
        if img_path and is_unreachable_screenshot(img_path):
            print(f"[SKIP] Unreachable: {result.get('url')}")
            continue

        ws[f"A{row}"] = result.get("url", "")
        ws[f"B{row}"] = result.get("ports", "")

        if img_path and os.path.isfile(img_path):
            img = ExcelImage(img_path)

            # Resize image
            img.width = 300
            img.height = 180

            # Adjust row height to match image
            ws.row_dimensions[row].height = 135

            # Embed image in column C
            ws.add_image(img, f"C{row}")

        row += 1

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 50  # URL
    ws.column_dimensions[get_column_letter(2)].width = 20  # Ports
    ws.column_dimensions[get_column_letter(3)].width = 45  # Screenshot

    wb.save(output_path)
    print(f"[✓] Excel saved with embedded images at: {output_path}")
